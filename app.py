import streamlit as st
import pandas as pd
import pdfplumber
import re
from docxtpl import DocxTemplate
from datetime import datetime, timezone, timedelta
import io
import os
import sys
import openpyxl
import openpyxl.utils  
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage, ImageChops
import fitz  # PyMuPDF
import numpy as np
import gc
import math
import zipfile
import time
from streamlit_sortables import sort_items

# ==============================================================================
# [공통 유틸리티 및 화면 설정]
# ==============================================================================
st.set_page_config(page_title="통합 양식 변환 및 검토 시스템", layout="wide")

# [CSS] 파일 업로더 레이아웃 및 업로드 목록 숨김 처리, UI 디테일 수정
st.markdown("""
    <style>
    /* 1. 파일 업로더 폭과 높이를 버튼과 안내 텍스트에 딱 맞게 슬림하게 축소 */
    [data-testid="stFileUploader"] { width: 100%; }
    [data-testid="stFileUploaderDropzone"] { 
        padding: 0.5rem 1rem !important; 
        min-height: auto !important; 
        display: flex; 
        flex-direction: column; 
        justify-content: center; 
        align-items: flex-start; 
    }
    [data-testid="stFileUploaderFileName"] { display: none; }
    [data-testid="stFileUploaderFileData"] { display: none; }
    div[data-testid="stHorizontalBlock"] div div div div { display: block !important; width: 100% !important; }
    
    /* 2. 모드 선택(Selectbox) 클릭 시 타자 입력 커서 숨김 및 글자 위아래 쏠림(잘림) 완벽 해결 */
    div[data-baseweb="select"] input {
        caret-color: transparent !important;
        cursor: pointer !important;
        position: absolute !important; /* 입력창을 문서 흐름에서 제외하여 글씨 밀림 원천 차단 */
        opacity: 0 !important;         /* 시각적으로 투명하게 처리 */
        height: 0 !important;
    }
    
    /* 3. 신버전/구버전 라디오 버튼 글씨 줄바꿈 방지 */
    div[role="radiogroup"] label {
        white-space: nowrap !important;
    }
    
    /* 4. 기타 양식(OTHERS) 체크박스 줄바꿈 방지 및 동일 간격 설정 */
    div[data-testid="stCheckbox"] label p {
        white-space: nowrap !important;
        overflow: hidden !important;
        text-overflow: ellipsis !important;
        line-height: 1.5 !important;
    }
    div[data-testid="stCheckbox"] {
        min-height: auto !important;
        padding: 0 !important;
    }
    div.row-widget.stCheckbox {
        margin-top: 0 !important;
        margin-bottom: 0.2rem !important; /* 항목 간 간격을 0.2rem으로 균일하게 통일 */
    }
    
    /* 5. 기타 양식 체크박스 선택 시 글씨 색상 빨간색으로 변경 */
    div[data-testid="stCheckbox"] label input[type="checkbox"]:checked ~ div p {
        color: red !important;
        font-weight: bold !important;
    }
    </style>
    """, unsafe_allow_html=True)


# ==============================================================================
# [패스워드 인증 로직 (완벽한 블러 및 클릭/조작 차단 적용)]
# ==============================================================================
if 'authenticated' not in st.session_state:
    st.session_state['authenticated'] = False

if not st.session_state['authenticated']:
    st.markdown("""
        <style>
        /* 1. 하위 UI를 완벽히 가리고 클릭을 차단하는 투명 방어막(Overlay) 생성 */
        #blur-overlay {
            position: fixed !important;
            top: 0 !important; left: 0 !important;
            width: 100vw !important; height: 100vh !important;
            background-color: rgba(255, 255, 255, 0.4) !important;
            backdrop-filter: blur(12px) !important;
            -webkit-backdrop-filter: blur(12px) !important;
            z-index: 999990 !important;
            pointer-events: auto !important; /* 클릭, 탭 등 모든 상호작용 여기서 차단 */
        }
        
        /* 2. 로그인 창(Form)만 방어막 뚫고 맨 위로 올라오게 설정 */
        [data-testid="stForm"] {
            position: fixed !important;
            top: 50% !important; left: 50% !important;
            transform: translate(-50%, -50%) !important;
            z-index: 999999 !important;
            background: white !important;
            padding: 2.5rem !important;
            border-radius: 15px !important;
            box-shadow: 0px 10px 40px rgba(0,0,0,0.3) !important;
            width: 350px !important;
            border: 1px solid #ddd !important;
        }
        
        /* 3. 사이드바와 상단 헤더가 방어막 위로 올라오지 못하도록 강제 블러 및 층위 하락 */
        [data-testid="stHeader"], [data-testid="stSidebar"] {
            z-index: 0 !important;
            filter: blur(12px) !important;
            pointer-events: none !important;
        }
        
        /* 4. 인증 전에 스크롤바 조작 원천 차단 */
        body {
            overflow: hidden !important;
        }
        </style>
        
        <div id="blur-overlay"></div>
    """, unsafe_allow_html=True)
    
    # 5. 정중앙에 띄워질 패스워드 입력 폼 생성 (st.stop을 안 쓰므로 하위 UI가 렌더되어 블러 효과가 나타남)
    with st.form("login_form"):
        st.markdown("<h3 style='text-align:center; margin-top:0;'>🔒 보안 인증</h3>", unsafe_allow_html=True)
        st.markdown("<p style='text-align:center; color:#666; font-size:14px; margin-bottom:20px;'>프로그램을 사용하시려면<br>패스워드를 입력해주세요.</p>", unsafe_allow_html=True)
        pwd = st.text_input("패스워드", type="password", label_visibility="collapsed", placeholder="패스워드 입력")
        submitted = st.form_submit_button("프로그램 실행", use_container_width=True)
        
        if submitted:
            if pwd == "a1234":
                st.session_state['authenticated'] = True
                st.rerun() # 인증 성공 시 화면을 재시작하여 방어막 제거
            else:
                st.error("❌ 패스워드가 올바르지 않습니다.")


# --- [종료 버튼 기능] ---
with st.sidebar:
    st.write("---")
    if st.button("❌ 프로그램 종료", type="primary"):
        st.warning("프로그램을 종료합니다. 창을 닫으셔도 됩니다.")
        time.sleep(1)
        os._exit(0) # 프로세스 강제 종료

def get_resource_path(relative_path):
    """실행 파일(exe) 내부나 일반 환경에서 절대 경로를 찾습니다."""
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# 세션 상태 초기화 (결과물 유지용)
SESSION_KEYS = [
    'spec_res', 'spec_fname',
    'allergy_res_83', 'allergy_res_26', 'allergy_fname_83', 'allergy_fname_26',
    'ifra_res', 'ifra_fname',
    'msds_res', # 리스트 형태: [{'fname': '...', 'data': ...}]
    'others_res', 'others_fname'
]
for k in SESSION_KEYS:
    if k not in st.session_state:
        st.session_state[k] = None if 'res' not in k or k == 'msds_res' else ""
if st.session_state['msds_res'] is None:
    st.session_state['msds_res'] = []


# ==============================================================================
# [파트 1: 통합 양식 변환기 내부 함수]
# ==============================================================================
def process_spec(pdf_file, product_name, mode):
    pdf_text = ""
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            extracted = page.extract_text()
            if extracted:
                pdf_text += extracted + "\n"
                
    context = {
        "PRODUCT": product_name,
        "COLOR": "PALE YELLOW TO YELLOW",
        "SG": "0.902 ~ 0.922",
        "RI": "1.466 ~ 1.476",
        "DATE": datetime.now().strftime("%d. %b. %Y").upper()
    }
    
    if mode == "CFF":
        color_match = re.search(r'COLOR\s*:(.*?)APPEARANCE\s*:', pdf_text, re.DOTALL | re.IGNORECASE)
        if color_match: context["COLOR"] = color_match.group(1).strip().upper()
        sg_match = re.search(r'SPECIFIC GRAVITY.*?\(\d+°C\)\s*:\s*([\d\.]+)\s*[±\+/-]\s*[\d\.]+', pdf_text, re.IGNORECASE)
        if sg_match:
            sg_base = float(sg_match.group(1))
            context["SG"] = f"{sg_base - 0.01:.3f} ~ {sg_base + 0.01:.3f}"
        ri_match = re.search(r'REFRACTIVE INDEX.*?\(\d+°C\)\s*:\s*([\d\.]+)\s*[±\+/-]\s*[\d\.]+', pdf_text, re.IGNORECASE)
        if ri_match:
            ri_base = float(ri_match.group(1))
            context["RI"] = f"{ri_base - 0.005:.3f} ~ {ri_base + 0.005:.3f}"
    elif mode == "HP":
        color_match = re.search(r'■\s*COLOR\s*:(.*?)■\s*APPEARANCE\s*:', pdf_text, re.DOTALL | re.IGNORECASE)
        if color_match: context["COLOR"] = color_match.group(1).strip().upper()
        sg_match = re.search(r'■\s*SPECIFIC GRAVITY.*?\:\s*([\d\.]+)\s*[±\+/-]\s*[\d\.]+', pdf_text, re.IGNORECASE)
        if sg_match:
            sg_base = float(sg_match.group(1))
            context["SG"] = f"{sg_base - 0.01:.3f} ~ {sg_base + 0.01:.3f}"
        ri_match = re.search(r'■\s*REFRACTIVE INDEX.*?\:\s*([\d\.]+)\s*[±\+/-]\s*[\d\.]+', pdf_text, re.IGNORECASE)
        if ri_match:
            ri_base = float(ri_match.group(1))
            context["RI"] = f"{ri_base - 0.005:.3f} ~ {ri_base + 0.005:.3f}"

    doc_path = get_resource_path("SPEC templates/spec.docx")
    doc = DocxTemplate(doc_path)
    doc.render(context)
    bio = io.BytesIO()
    doc.save(bio)
    bio.seek(0)
    return bio, f"{product_name} SPEC.docx"


def extract_cas(text):
    if pd.isna(text): return []
    clean_text = str(text).replace('/', ' ').replace('\n', ' ').replace('\r', ' ')
    return re.findall(r'\d{2,7}-\d{2}-\d', clean_text)

def logic_cff_83(input_df, template_path, customer_name, product_name):
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    for row in ws.iter_rows(min_col=3, max_col=3, min_row=1):
        for cell in row:
            if not isinstance(cell, MergedCell) and str(cell.value).startswith('='): cell.value = None
    if "Sheet2" in wb.sheetnames: del wb["Sheet2"]
    source_data = {}
    for idx, row in input_df.iterrows():
        cas_text = row.iloc[5] if len(row) > 5 else None
        val = row.iloc[11] if len(row) > 11 else None
        for cas in extract_cas(cas_text): source_data[cas] = val
    for r in range(1, ws.max_row + 1):
        template_cas_text = ws.cell(row=r, column=2).value
        if template_cas_text:
            for t_cas in extract_cas(template_cas_text):
                if t_cas in source_data:
                    target = ws.cell(row=r, column=3)
                    if not isinstance(target, MergedCell): target.value = source_data[t_cas]
                    break 
    ws['B9'] = customer_name
    ws['B10'] = product_name
    ws['E10'] = datetime.now().strftime("%Y-%m-%d")
    return wb

def logic_cff_26(input_df, template_path, customer_name, product_name):
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    for row in ws.iter_rows(min_col=3, max_col=3, min_row=18, max_row=43):
        for cell in row: 
            if not isinstance(cell, MergedCell): cell.value = None
    source_data = {}
    for idx, row in input_df.iterrows():
        cas_text = row.iloc[5] if len(row) > 5 else None
        val = row.iloc[11] if len(row) > 11 else None
        for cas in extract_cas(cas_text): source_data[cas] = val
        
    for r in range(1, ws.max_row + 1):
        template_cas_text = ws.cell(row=r, column=2).value
        if template_cas_text:
            for t_cas in extract_cas(template_cas_text):
                if t_cas in source_data:
                    target = ws.cell(row=r, column=3)
                    if not isinstance(target, MergedCell): target.value = source_data[t_cas]
                    break
                    
    # C18:43 범위에 숫자가 하나라도 있는지 검사
    has_num = False
    for r in range(18, 44):
        val = ws.cell(row=r, column=3).value
        if val is not None and str(val).strip() != "":
            if any(char.isdigit() for char in str(val)):
                has_num = True
                break

    # 숫자가 하나도 없다면 C18:43 범위를 모두 0으로 덮어씀
    if not has_num:
        for r in range(18, 44):
            target = ws.cell(row=r, column=3)
            try:
                if not isinstance(target, MergedCell): target.value = 0
            except: pass

    ws['B11'] = customer_name; ws['B12'] = product_name; ws['E13'] = datetime.now().strftime("%Y-%m-%d")
    align_center = Alignment(horizontal='center', vertical='center')
    for row in ws.iter_rows(min_col=3, max_col=6, min_row=18, max_row=43):
        for cell in row: 
            if not isinstance(cell, MergedCell): cell.alignment = align_center
    return wb

def logic_hp_83(input_df, template_path, customer_name, product_name):
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    for row in ws.iter_rows(min_col=3, max_col=3, min_row=1):
        for cell in row:
            if not isinstance(cell, MergedCell) and str(cell.value).startswith('='): cell.value = None
    if "Sheet2" in wb.sheetnames: del wb["Sheet2"]
    source_data = {}
    for idx, row in input_df.iterrows():
        cas_text = row.iloc[1] if len(row) > 1 else None
        val = row.iloc[2] if len(row) > 2 else None
        for cas in extract_cas(cas_text): source_data[cas] = val
    for r in range(1, ws.max_row + 1):
        template_cas_text = ws.cell(row=r, column=2).value
        if template_cas_text:
            for t_cas in extract_cas(template_cas_text):
                if t_cas in source_data:
                    val_to_insert = source_data[t_cas]
                    if pd.notna(val_to_insert) and str(val_to_insert).strip() not in ['0', '0.0']:
                        target = ws.cell(row=r, column=3)
                        if not isinstance(target, MergedCell): target.value = val_to_insert
                    break 
    ws['B9'] = customer_name; ws['B10'] = product_name; ws['E10'] = datetime.now().strftime("%Y-%m-%d")
    return wb

def logic_hp_26(input_df, template_path, customer_name, product_name):
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    for row in ws.iter_rows(min_col=3, max_col=3, min_row=18, max_row=43):
        for cell in row: 
            if not isinstance(cell, MergedCell): cell.value = None
    source_data = {}
    for idx, row in input_df.iterrows():
        cas_text = row.iloc[1] if len(row) > 1 else None
        val = row.iloc[2] if len(row) > 2 else None
        for cas in extract_cas(cas_text): source_data[cas] = val
        
    for r in range(1, ws.max_row + 1):
        template_cas_text = ws.cell(row=r, column=2).value
        if template_cas_text:
            for t_cas in extract_cas(template_cas_text):
                if t_cas in source_data:
                    val_to_insert = source_data[t_cas]
                    if pd.notna(val_to_insert) and str(val_to_insert).strip() not in ['0', '0.0']:
                        target = ws.cell(row=r, column=3)
                        if not isinstance(target, MergedCell): target.value = val_to_insert
                    break
                    
    # C18:43 범위에 숫자가 하나라도 있는지 검사
    has_num = False
    for r in range(18, 44):
        val = ws.cell(row=r, column=3).value
        if val is not None and str(val).strip() != "":
            if any(char.isdigit() for char in str(val)):
                has_num = True
                break

    # 숫자가 하나도 없다면 C18:43 범위를 모두 0으로 덮어씀
    if not has_num:
        for r in range(18, 44):
            target = ws.cell(row=r, column=3)
            try:
                if not isinstance(target, MergedCell): target.value = 0
            except: pass

    ws['B11'] = customer_name; ws['B12'] = product_name; ws['E13'] = datetime.now().strftime("%Y-%m-%d")
    align_center = Alignment(horizontal='center', vertical='center')
    for row in ws.iter_rows(min_col=3, max_col=6, min_row=18, max_row=43):
        for cell in row: 
            if not isinstance(cell, MergedCell): cell.alignment = align_center
    return wb

def to_excel(data):
    output = io.BytesIO()
    if isinstance(data, pd.DataFrame):
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            data.to_excel(writer, index=False, sheet_name='Sheet1')
    else:
        data.save(output)
    return output.getvalue()

def extract_text_between_ifra(text, start_keyword, end_keyword=None):
    def flexible_escape(kw):
        escaped = re.escape(kw).replace(r'\ ', r'\s+').replace(' ', r'\s+')
        escaped = escaped.replace(r'\.', r'\s*\.?\s*').replace(r'\*', r'\s*\*\s*')
        return escaped
    start_pattern = flexible_escape(start_keyword)
    if end_keyword:
        end_pattern = flexible_escape(end_keyword)
        match = re.search(f"{start_pattern}(.*?){end_pattern}", text, re.DOTALL | re.IGNORECASE)
    else:
        match = re.search(f"{start_pattern}(.*)", text, re.IGNORECASE)
    if match: return process_value_ifra(match.group(1).strip())
    return "Not Permitted"

def process_value_ifra(val_str):
    if isinstance(val_str, (int, float)): val_str = str(val_str)
    if not val_str: return "Not Permitted"
    val_lower = val_str.lower()
    if "not" in val_lower and "permitted" in val_lower: return "Not Permitted"
    if "not" in val_lower and "restricted" in val_lower: return "Not Restricted"
    num_match = re.search(r'\d+\.?\d*', val_str)
    if not num_match: return "Not Permitted"
    try:
        clean_str = num_match.group(0)
        val_float = float(clean_str)
        if val_float == 0.0: return "Not Permitted"
        s = str(val_float)
        if '.' in s:
            int_part, dec_part = s.split('.')
            dec_part = dec_part[:2].ljust(2, '0')
            return f"{int_part}.{dec_part}%"
        return f"{s}.00%"
    except ValueError: return "Not Permitted"

def process_ifra(pdf_file, customer_name, product_name, mode):
    full_text = ""
    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages: full_text += page.extract_text() + "\n"
    if mode == "CFF":
        context = {
            "CUSTOMER": customer_name, "PRODUCT": product_name,
            "CATEGORY1": extract_text_between_ifra(full_text, "Category 1", "Category 2"),
            "CATEGORY2": extract_text_between_ifra(full_text, "Category 2", "Category 3"),
            "CATEGORY3": extract_text_between_ifra(full_text, "Category 3", "Category 4"),
            "CATEGORY4": extract_text_between_ifra(full_text, "Category 4", "Category 5.A"),
            "CATEGORY5_A": extract_text_between_ifra(full_text, "Category 5.A", "Category 5.B"),
            "CATEGORY5_B": extract_text_between_ifra(full_text, "Category 5.B", "Category 5.C"),
            "CATEGORY5_C": extract_text_between_ifra(full_text, "Category 5.C", "Category 5.D"),
            "CATEGORY5_D": extract_text_between_ifra(full_text, "Category 5.D", "Category 6"),
            "CATEGORY6": extract_text_between_ifra(full_text, "Category 6", "Category 7.A"),
            "CATEGORY7_A": extract_text_between_ifra(full_text, "Category 7.A", "Category 7.B"),
            "CATEGORY7_B": extract_text_between_ifra(full_text, "Category 7.B", "Category 8"),
            "CATEGORY8": extract_text_between_ifra(full_text, "Category 8", "Category 9"),
            "CATEGORY9": extract_text_between_ifra(full_text, "Category 9", "Category 10.A"),
            "CATEGORY10_A": extract_text_between_ifra(full_text, "Category 10.A", "Category 10.B"),
            "CATEGORY10_B": extract_text_between_ifra(full_text, "Category 10.B", "Category 11.A"),
            "CATEGORY11_A": extract_text_between_ifra(full_text, "Category 11.A", "Category 11.B"),
            "CATEGORY11_B": extract_text_between_ifra(full_text, "Category 11.B", "Category 12"),
            "CATEGORY12": extract_text_between_ifra(full_text, "Category 12", None)
        }
    elif mode == "HP":
        context = {
            "CUSTOMER": customer_name, "PRODUCT": product_name,
            "CATEGORY1": extract_text_between_ifra(full_text, "Category 1*", "Category 2"),
            "CATEGORY2": extract_text_between_ifra(full_text, "Category 2", "Category 3"),
            "CATEGORY3": extract_text_between_ifra(full_text, "Category 3", "Category 4"),
            "CATEGORY4": extract_text_between_ifra(full_text, "Category 4", "Category 5.A"),
            "CATEGORY5_A": extract_text_between_ifra(full_text, "Category 5.A", "Category 5.B"),
            "CATEGORY5_B": extract_text_between_ifra(full_text, "Category 5.B", "Category 5.C"),
            "CATEGORY5_C": extract_text_between_ifra(full_text, "Category 5.C", "Category 5.D"),
            "CATEGORY5_D": extract_text_between_ifra(full_text, "Category 5.D", "Category 6*"),
            "CATEGORY6": extract_text_between_ifra(full_text, "Category 6*", "Category 7.A"),
            "CATEGORY7_A": extract_text_between_ifra(full_text, "Category 7.A", "Category 7.B"),
            "CATEGORY7_B": extract_text_between_ifra(full_text, "Category 7.B", "Category 8"),
            "CATEGORY8": extract_text_between_ifra(full_text, "Category 8", "Category 9"),
            "CATEGORY9": extract_text_between_ifra(full_text, "Category 9", "Category 10.A"),
            "CATEGORY10_A": extract_text_between_ifra(full_text, "Category 10.A", "Category 10.B"),
            "CATEGORY10_B": extract_text_between_ifra(full_text, "Category 10.B", "Category 11.A"),
            "CATEGORY11_A": extract_text_between_ifra(full_text, "Category 11.A", "Category 11.B"),
            "CATEGORY11_B": extract_text_between_ifra(full_text, "Category 11.B", "Category 12"),
            "CATEGORY12": extract_text_between_ifra(full_text, "Category 12", None)
        }
    template_path = get_resource_path("IFRA templates/IFRA.docx")
    doc = DocxTemplate(template_path)
    doc.render(context)
    output_io = io.BytesIO()
    doc.save(output_io)
    output_io.seek(0)
    return output_io, f"{product_name} IFRA 51TH.docx"

TARGET_23_CAS = {
    "127-51-5", "122-40-7", "101-85-9", "105-13-5", "100-51-6",
    "120-51-4", "103-41-3", "118-58-1", "104-55-2", "104-54-1",
    "5392-40-5", "106-22-9", "91-64-5", "5989-27-5", "97-53-0",
    "4602-84-0", "106-24-1", "101-86-0", "107-75-5", "97-54-1",
    "78-70-6", "31906-04-4", "80-54-6", "111-12-6", "90028-68-5", "90028-67-4"
}

def convert_xls_to_xlsx(uploaded_file):
    if uploaded_file.name.lower().endswith('.xls'):
        df_dict = pd.read_excel(uploaded_file, sheet_name=None, engine='xlrd')
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for sheet_name, df in df_dict.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        output.seek(0)
        return output
    return uploaded_file

def get_cas_set(cas_val):
    if not cas_val: return frozenset()
    cas_list = re.findall(r'\d+-\d+-\d+', str(cas_val))
    return frozenset(cas.strip() for cas in cas_list)

def handle_upload(col, label, key):
    with col:
        st.subheader(label)
        uploaded = st.file_uploader(f"{label} 선택", type=["xlsx", "xls"], accept_multiple_files=True, key=key)
        sorted_list = []
        if uploaded:
            display_items = [f"↕ {i+1}. {f.name}" for i, f in enumerate(uploaded)]
            sorted_names = sort_items(display_items, direction="vertical", key=f"sort_v3_{key}_{len(uploaded)}")
            for name in sorted_names:
                try:
                    orig_name = name.split(". ", 1)[1]
                    matched_file = next((f for f in uploaded if f.name == orig_name), None)
                    if matched_file:
                        sorted_list.append(matched_file)
                except (IndexError, StopIteration):
                    continue
        return sorted_list

def extract_data(file_raw, is_23=False, is_83=False):
    f = convert_xls_to_xlsx(file_raw)
    wb = load_workbook(f, data_only=True)
    ws = wb.worksheets[0]
    name_upper = file_raw.name.upper()
    data_map = {}
    product_name = "알 수 없음"
    
    def clean_val(v):
        if v is None or str(v).strip() == "-": return 0.0
        try: return float(v)
        except: return 0.0

    val_a1 = str(ws.cell(row=1, column=1).value or "").strip()
    val_b1 = str(ws.cell(row=1, column=2).value or "").strip()
    
    # 1. CFF 로직 (기존 유지)
    if val_a1 == "성분코드" and val_b1 == "성분국문명":
        product_name = file_raw.name
        empty_count = 0
        for r in range(2, ws.max_row + 1):
            cas_raw = ws.cell(row=r, column=6).value
            if cas_raw is None or str(cas_raw).strip() == "":
                empty_count += 1
                if empty_count >= 10: break
            else:
                empty_count = 0
            
            c, v = get_cas_set(cas_raw), clean_val(ws.cell(row=r, column=8).value)
            if c and v != 0: 
                data_map[c] = {"n": ws.cell(row=r, column=2).value, "v": v}
                
    # 2. HP 통합 로직 (포괄적 키워드 탐색 적용)
    elif "HP" in name_upper:
        product_name = ws.cell(row=10, column=2).value or file_raw.name
        
        # 열 인덱스를 찾을 변수
        col_name, col_cas, col_val = None, None, None
        
        # [핵심 수정] 1행에만 헤더가 있는게 아니므로 셀 전체를 돌며 열 번호를 찾음
        for r in range(1, min(ws.max_row + 1, 100)): # 상위 100행까지 스캔
            for c in range(1, ws.max_column + 1):
                # 띄어쓰기, 줄바꿈 등 모든 공백을 없애서 텍스트 매칭 신뢰도 100% 보장
                cell_text = str(ws.cell(row=r, column=c).value or "").upper()
                cell_text_clean = cell_text.replace('\n', '').replace('\r', '').replace('\xa0', '').replace(' ', '')
                
                if not cell_text_clean:
                    continue
                    
                # 물질명 찾는 키워드 (Benzyl alcohol 무조건 포함된다고 하셨으므로)
                if col_name is None and "BENZYLALCOHOL" in cell_text_clean:
                    col_name = c
                # CAS NO 찾는 키워드
                if col_cas is None and "CAS" in cell_text_clean:
                    col_cas = c
                # 수치 찾는 키워드 (Total in Fragrance Oil(%) 공백 제거본)
                if col_val is None and "TOTALINFRAGRANCEOIL(%)" in cell_text_clean:
                    col_val = c
            
            # 3개의 열을 모두 찾았다면 더 이상 불필요한 스캔 중단
            if col_name and col_cas and col_val:
                break
                
        # 만약 양식이 너무 달라서 아예 키워드를 못 찾았을 때를 대비한 기본값(Fallback)
        if not col_name: col_name = 1
        if not col_cas: col_cas = 2
        if not col_val: col_val = 3

        # 데이터 매핑 (한 행 안에서 찾은 열 번호들을 기준으로 동일 물질 정보 묶기)
        for r in range(1, ws.max_row + 1):
            name = ws.cell(row=r, column=col_name).value
            cas = ws.cell(row=r, column=col_cas).value
            val = ws.cell(row=r, column=col_val).value
            
            c_set = get_cas_set(cas)
            v = clean_val(val)
            
            if c_set and v != 0:
                data_map[c_set] = {"n": str(name).strip() if name else "지정성분", "v": v}

    # 3. 기타(is_83, is_23) 로직 (기존 유지)
    elif is_83:
        product_name = ws.cell(row=10, column=2).value
        for r in range(1, ws.max_row + 1):
            cas_raw_b, cas_raw_c = ws.cell(row=r, column=2).value, ws.cell(row=r, column=3).value
            c = get_cas_set(cas_raw_b) if get_cas_set(cas_raw_b) else get_cas_set(cas_raw_c)
            v = clean_val(ws.cell(row=r, column=4).value if (not get_cas_set(cas_raw_b) and get_cas_set(cas_raw_c)) else ws.cell(row=r, column=3).value)
            if c and v != 0: data_map[c] = {"n": ws.cell(row=r, column=1).value, "v": v}
            
    elif is_23:
        product_name = ws.cell(row=12, column=2).value
        for r in range(18, 44):
            c, v = get_cas_set(ws.cell(row=r, column=2).value), clean_val(ws.cell(row=r, column=3).value)
            if c and v != 0: data_map[c] = {"n": ws.cell(row=r, column=1).value or "지정성분", "v": v}
            
    else:
        product_name = ws.cell(row=7, column=4).value
        for r in range(13, ws.max_row + 1):
            cas_raw = ws.cell(row=r, column=6).value
            c, v = get_cas_set(cas_raw), clean_val(ws.cell(row=r, column=12).value)
            if c and v != 0: data_map[c] = {"n": ws.cell(row=r, column=2).value, "v": v}
    
    wb.close()
    return str(product_name).strip() if product_name else file_raw.name, data_map

def process_others(customer_name, product_name, selected_files):
    """
    선택된 기타 양식(.docx) 템플릿들에 고객사, 제품명, 날짜를 입력하고
    하나의 ZIP 파일로 압축하여 반환하는 함수입니다.
    """
    try:
        zip_buffer = io.BytesIO()
        template_dir = get_resource_path("OTHERS templates")
        
        # ZIP 파일 생성 준비
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            for file_name in selected_files:
                template_path = os.path.join(template_dir, file_name)
                
                # 워드 템플릿 열기
                doc = DocxTemplate(template_path)
                
                # 템플릿에 들어갈 데이터 ({{CUSTOMER}}, {{PRODUCT}}, {{DATE}} 치환)
                context = {
                    "CUSTOMER": customer_name,
                    "PRODUCT": product_name,
                    "DATE": datetime.now().strftime("%d. %b. %Y").upper()
                }
                
                # 데이터 렌더링
                doc.render(context)
                
                # 변환된 문서를 메모리에 임시 저장
                doc_io = io.BytesIO()
                doc.save(doc_io)
                doc_io.seek(0)
                
                # 제품명이 포함된 새로운 파일명으로 ZIP 안에 추가
                output_name = f"{product_name} {file_name}"
                zip_file.writestr(output_name, doc_io.getvalue())
        
        zip_buffer.seek(0)
        return zip_buffer, f"{product_name}_OTHERS.zip"
        
    except Exception as e:
        return None, f"기타 양식 처리 중 오류 발생: {e}"

# ==============================================================================
# [UI 레이아웃 구성: 파트 1 (통합 양식 변환기)]
# ==============================================================================
st.title("📄 아로마준 통합 양식 변환기")

# --- 공통 정보 및 일괄 변환 ---
st.subheader("공통 정보 입력")
col_top1, col_top2, col_top_mode, col_top3 = st.columns([2, 2, 2, 1])
with col_top1:
    global_customer = st.text_input("고객사명 (CUSTOMER)")
with col_top2:
    global_product = st.text_input("제품명 (PRODUCT)")
with col_top_mode:
    global_mode = st.selectbox("모드 선택", ["CFF(K)", "CFF(E)", "HP(K)", "HP(E)"])
with col_top3:
    st.write("")
    batch_run = st.button("🌟 일괄 변환 실행", use_container_width=True)

# 통합 모드를 기준으로 SPEC, ALLERGY, IFRA용 기본 모드(CFF or HP) 판별
base_mode = "CFF" if "CFF" in global_mode else "HP"

st.divider()

# --- Section 1: SPEC ---
col1_1, col1_2, col1_3 = st.columns(3)
with col1_1:
    st.subheader("1. SPEC 양식 변환")
    spec_up = st.file_uploader("원본 PDF 업로드", type=["pdf"], key="spec_up")
with col1_2:
    st.subheader(" ")
    if st.button("SPEC 변환", use_container_width=True):
        if not spec_up or not global_product: st.warning("원본 파일과 제품명을 입력해주세요.")
        else:
            with st.spinner("SPEC 변환 중..."):
                try:
                    res, fname = process_spec(spec_up, global_product, base_mode)
                    st.session_state['spec_res'] = res.getvalue()
                    st.session_state['spec_fname'] = fname
                    st.success("변환 성공!")
                except Exception as e: st.error(f"오류: {e}")
with col1_3:
    st.subheader("결과물 다운로드")
    if st.session_state['spec_res']:
        c_n, c_b = st.columns([3, 1])
        c_n.write(f"📄 {st.session_state['spec_fname']}")
        c_b.download_button("다운로드", data=st.session_state['spec_res'], file_name=st.session_state['spec_fname'], mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document", key="dl_spec")

st.divider()

# --- Section 2: ALLERGY ---
col2_1, col2_2, col2_3 = st.columns(3)
with col2_1:
    st.subheader("2. ALLERGY 양식 변환")
    allergy_up = st.file_uploader("원본 Excel 업로드", type=['xlsx', 'xls'], key="allergy_up")
with col2_2:
    st.subheader(" ")
    if st.button("ALLERGY 변환", use_container_width=True):
        if not allergy_up or not global_customer or not global_product: st.warning("원본 파일, 고객사명, 제품명을 모두 입력해주세요.")
        else:
            with st.spinner("ALLERGY 변환 중..."):
                try:
                    input_df = pd.read_excel(allergy_up)
                    base_path = get_resource_path("ALLERGY templates")
                    if base_mode == "CFF":
                        res_83 = logic_cff_83(input_df, os.path.join(base_path, "83 CFF.xlsx"), global_customer, global_product)
                        res_26 = logic_cff_26(input_df, os.path.join(base_path, "26 통합.xlsx"), global_customer, global_product)
                    else:
                        res_83 = logic_hp_83(input_df, os.path.join(base_path, "83 HP.xlsx"), global_customer, global_product)
                        res_26 = logic_hp_26(input_df, os.path.join(base_path, "26 통합.xlsx"), global_customer, global_product)
                    st.session_state['allergy_res_83'] = to_excel(res_83)
                    st.session_state['allergy_res_26'] = to_excel(res_26)
                    st.session_state['allergy_fname_83'] = f"83 ALLERGENS {global_product}.xlsx"
                    st.session_state['allergy_fname_26'] = f"ALLERGEN {global_product}.xlsx"
                    st.success("변환 성공!")
                except Exception as e: st.error(f"오류: {e}")
with col2_3:
    st.subheader("결과물 다운로드")
    if st.session_state['allergy_res_83'] and st.session_state['allergy_res_26']:
        c1_n, c1_b = st.columns([3, 1])
        c1_n.write(f"📄 {st.session_state['allergy_fname_83']}")
        c1_b.download_button("다운로드", data=st.session_state['allergy_res_83'], file_name=st.session_state['allergy_fname_83'], mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_al_83")
        c2_n, c2_b = st.columns([3, 1])
        c2_n.write(f"📄 {st.session_state['allergy_fname_26']}")
        c2_b.download_button("다운로드", data=st.session_state['allergy_res_26'], file_name=st.session_state['allergy_fname_26'], mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_al_26")

st.divider()

# --- Section 3: IFRA ---
col3_1, col3_2, col3_3 = st.columns(3)
with col3_1:
    st.subheader("3. IFRA 양식 변환")
    ifra_up = st.file_uploader("원본 PDF 업로드", type=["pdf"], key="ifra_up")
with col3_2:
    st.subheader(" ")
    if st.button("IFRA 변환", use_container_width=True):
        if not ifra_up or not global_customer or not global_product: st.warning("원본 파일, 고객사명, 제품명을 모두 입력해주세요.")
        else:
            with st.spinner("IFRA 변환 중..."):
                try:
                    res, fname = process_ifra(ifra_up, global_customer, global_product, base_mode)
                    st.session_state['ifra_res'] = res.getvalue()
                    st.session_state['ifra_fname'] = fname
                    st.success("변환 성공!")
                except Exception as e: st.error(f"오류: {e}")
with col3_3:
    st.subheader("결과물 다운로드")
    if st.session_state['ifra_res']:
        c_n, c_b = st.columns([3, 1])
        c_n.write(f"📄 {st.session_state['ifra_fname']}")
        c_b.download_button("다운로드", data=st.session_state['ifra_res'], file_name=st.session_state['ifra_fname'], mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document", key="dl_ifra")

st.divider()

# --- Section 4: MSDS ---
# (코드 간결화를 위해 MSDS 부분은 누락하지 말라는 요청에 따라 생략하지 않고 그대로 이어집니다 - 사용자의 원래 코드에는 생략되어 있었음)
col4_1, col4_2, col4_3 = st.columns(3)
with col4_1:
    st.subheader("4. MSDS 양식 변환")
    msds_up = st.file_uploader("원본 PDF 업로드", type=["pdf"], accept_multiple_files=True, key="msds_up")
with col4_2:
    st.subheader(" ")
    msds_ri = ""
    msds_kor_file = None
    msds_kor_ver = "신버전"
    
    if "HP" in global_mode:
        msds_ri = st.text_input("굴절률 입력", key="msds_ri")
    if "E" in global_mode:
        st.info("💡 영문 양식 생성 시 국문 파일 첨부")
        msds_kor_file = st.file_uploader("국문 엑셀 파일", type="xlsx", key="msds_kor_file")
        msds_kor_ver = st.radio("국문 양식 버전", ["신버전", "구버전"], horizontal=True, key="msds_kor_ver")

    if st.button("MSDS 변환", use_container_width=True):
        if not msds_up or not global_customer or not global_product: st.warning("고객사명, 제품명, 원본 파일을 모두 입력해주세요.")
        else:
            with st.spinner("MSDS 변환 중..."):
                res_dict = process_msds(msds_up, global_product, global_mode, msds_ri, msds_kor_file, msds_kor_ver)
                if "error" in res_dict:
                    st.error(res_dict["error"])
                else:
                    st.session_state['msds_res'] = [{'fname': f, 'data': res_dict["data"][f]} for f in res_dict["files"]]
                    st.success("변환 성공!")
with col4_3:
    st.subheader("결과물 다운로드")
    if st.session_state['msds_res']:
        for i, item in enumerate(st.session_state['msds_res']):
            c_n, c_b = st.columns([3, 1])
            c_n.write(f"📄 {item['fname']}")
            c_b.download_button("다운로드", data=item['data'], file_name=item['fname'], mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=f"dl_msds_{i}")

st.divider()

# --- Section 5: OTHERS ---
col5_1, col5_2, col5_3 = st.columns(3)
with col5_1:
    st.subheader("5. 기타(원본 불필요) 양식 변환")
    st.info("이 양식은 원본 파일 첨부가 필요 없으며, 입력한 고객사명과 제품명만 참고하여 선택한 파일들이 ZIP 형식으로 일괄 생성됩니다.")
    
    template_dir = get_resource_path("OTHERS templates")
    available_others = []
    if os.path.exists(template_dir):
        available_others = sorted([f for f in os.listdir(template_dir) if f.endswith(".docx") and not f.startswith("~")])
    
    selected_others = []
    if available_others:
        for i, f in enumerate(available_others):
            if st.checkbox(f.replace(".docx", ""), key=f"chk_other_{i}"):
                selected_others.append(f)
    else:
        st.warning("OTHERS templates 폴더에 변환 가능한 파일이 없습니다.")

with col5_2:
    st.subheader(" ")
    if st.button("기타 변환", use_container_width=True):
        if not global_customer or not global_product: st.warning("상단의 고객사명과 제품명을 모두 입력해주세요.")
        elif not selected_others: st.warning("변환할 파일을 하나 이상 체크해주세요.")
        else:
            with st.spinner("기타 양식 변환 중..."):
                res, info = process_others(global_customer, global_product, selected_others)
                if res:
                    st.session_state['others_res'] = res.getvalue()
                    st.session_state['others_fname'] = info
                    st.success("변환 성공!")
                else:
                    st.error(info)
with col5_3:
    st.subheader("결과물 다운로드")
    if st.session_state['others_res']:
        c_n, c_b = st.columns([3, 1])
        c_n.write(f"📦 {st.session_state['others_fname']}")
        c_b.download_button("다운로드", data=st.session_state['others_res'], file_name=st.session_state['others_fname'], mime="application/zip", key="dl_others_zip")

# ==============================================================================
# [일괄 변환 트리거]
# ==============================================================================
if batch_run:
    st.write("---")
    if not global_customer or not global_product:
        st.error("❗ 일괄 변환을 위해 상단의 [고객사명]과 [제품명]을 먼저 입력해주세요.")
    else:
        with st.spinner("🌟 일괄 변환을 진행 중입니다... 잠시만 기다려주세요."):
            # 1. SPEC
            if spec_up:
                try:
                    res, fname = process_spec(spec_up, global_product, base_mode)
                    st.session_state['spec_res'] = res.getvalue(); st.session_state['spec_fname'] = fname
                except Exception as e: st.error(f"SPEC 일괄 변환 오류: {e}")
            
            # 2. ALLERGY
            if allergy_up:
                try:
                    input_df = pd.read_excel(allergy_up)
                    base_path = get_resource_path("ALLERGY templates")
                    if base_mode == "CFF":
                        r83 = logic_cff_83(input_df, os.path.join(base_path, "83 CFF.xlsx"), global_customer, global_product)
                        r26 = logic_cff_26(input_df, os.path.join(base_path, "26 통합.xlsx"), global_customer, global_product)
                    else:
                        r83 = logic_hp_83(input_df, os.path.join(base_path, "83 HP.xlsx"), global_customer, global_product)
                        r26 = logic_hp_26(input_df, os.path.join(base_path, "26 통합.xlsx"), global_customer, global_product)
                    st.session_state['allergy_res_83'] = to_excel(r83)
                    st.session_state['allergy_res_26'] = to_excel(r26)
                    st.session_state['allergy_fname_83'] = f"83 ALLERGENS {global_product}.xlsx"
                    st.session_state['allergy_fname_26'] = f"ALLERGEN {global_product}.xlsx"
                except Exception as e: st.error(f"ALLERGY 일괄 변환 오류: {e}")

            # 3. IFRA
            if ifra_up:
                try:
                    res, fname = process_ifra(ifra_up, global_customer, global_product, base_mode)
                    st.session_state['ifra_res'] = res.getvalue(); st.session_state['ifra_fname'] = fname
                except Exception as e: st.error(f"IFRA 일괄 변환 오류: {e}")
                
            # 4. MSDS
            if msds_up:
                res_dict = process_msds(msds_up, global_product, global_mode, msds_ri, msds_kor_file, msds_kor_ver)
                if "error" in res_dict: st.error(f"MSDS 일괄 변환 오류: {res_dict['error']}")
                else: st.session_state['msds_res'] = [{'fname': f, 'data': res_dict["data"][f]} for f in res_dict["files"]]

            # 5. OTHERS (체크박스에서 선택된 파일만 변환)
            if selected_others:
                res, info = process_others(global_customer, global_product, selected_others)
                if res:
                    st.session_state['others_res'] = res.getvalue()
                    st.session_state['others_fname'] = info
                else: st.error(f"OTHERS 일괄 변환 오류: {info}")
            
            st.success("✅ 파일이 첨부되거나 선택된 모든 항목의 일괄 변환이 완료되었습니다. 각 섹션의 우측에서 결과를 다운로드하세요!")


# ==============================================================================
# 굵은 시각적 구분선
# ==============================================================================
st.markdown("<br><br><br><hr style='border: 3px solid #ddd;'><br><br>", unsafe_allow_html=True)


# ==============================================================================
# [UI 레이아웃 구성: 파트 2 (알러지 자료 통합 검토 시스템)]
# ==============================================================================
st.title("📄 ALLERGENS 자료 통합 검토 시스템(HP/CFF)")

mode_review = st.radio("검토 방식 선택", ["원본 vs 83알러지", "원본 vs 26알러지", "83알러지 vs 26알러지", "원본 vs 83알러지 vs 26알러지"], horizontal=True, key="review_mode")
st.info("파일들을 **동일한 순번**으로 배치하세요. 동일 순번끼리 매칭되어 검토합니다.")
st.markdown("---")

files_A, files_B, files_C = [], [], []
if mode_review == "원본 vs 83알러지 vs 26알러지":
    col1, col2, col3 = st.columns(3)
    cols = [col1, col2, col3]
    labels = ["원본", "83알러지", "26알러지"]
else:
    col1, col2 = st.columns(2)
    cols = [col1, col2]
    labels = mode_review.split(" vs ")

files_A = handle_upload(cols[0], labels[0], "upload_A")
files_B = handle_upload(cols[1], labels[1], "upload_B")
if mode_review == "원본 vs 83알러지 vs 26알러지":
    files_C = handle_upload(cols[2], labels[2], "upload_C")

st.markdown("---")

# 4. 검증 로직 및 결과 출력
ready = files_A and files_B
if mode_review == "원본 vs 83알러지 vs 26알러지": ready = ready and files_C

if ready:
    num_pairs = min(len(files_A), len(files_B), len(files_C)) if (mode_review == "원본 vs 83알러지 vs 26알러지") else min(len(files_A), len(files_B))
    
    for idx in range(num_pairs):
        try:
            p_name_1, m1 = extract_data(files_A[idx], is_23=("26알러지" in labels[0]), is_83=("83알러지" in labels[0]))
            p_name_2, m2 = extract_data(files_B[idx], is_23=("26알러지" in labels[1]), is_83=("83알러지" in labels[1]))
            
            m3 = None
            p_name_3 = None
            if mode_review == "원본 vs 83알러지 vs 26알러지":
                p_name_3, m3 = extract_data(files_C[idx], is_23=True)

            display_p_name = p_name_2 if "알러지" in labels[1] else p_name_1
            if mode_review == "원본 vs 83알러지 vs 26알러지":
                display_p_name = p_name_2

            if "26알러지" in mode_review:
                m1 = {cas: d for cas, d in m1.items() if not cas.isdisjoint(TARGET_23_CAS)}
                m2 = {cas: d for cas, d in m2.items() if not cas.isdisjoint(TARGET_23_CAS)}
                if m3: m3 = {cas: d for cas, d in m3.items() if not cas.isdisjoint(TARGET_23_CAS)}

            all_cas_keys = set(m1.keys()) | set(m2.keys())
            if m3: all_cas_keys |= set(m3.keys())

            # [수정됨] 중복 행 발생 방지: 겹치는 CAS 번호들을 하나의 교집합 그룹으로 완벽히 병합하는 로직
            consolidated_cas_sets = []
            for cas in all_cas_keys:
                overlapping_indices = [i for i, c in enumerate(consolidated_cas_sets) if not cas.isdisjoint(c)]
                if not overlapping_indices:
                    consolidated_cas_sets.append(cas)
                else:
                    merged_set = set(cas)
                    for i in sorted(overlapping_indices, reverse=True):
                        merged_set |= consolidated_cas_sets.pop(i)
                    consolidated_cas_sets.append(frozenset(merged_set))

            rows, mismatch = [], 0
            
            for cas in consolidated_cas_sets:
                v1_data = next((m1[c] for c in m1 if not cas.isdisjoint(c)), None)
                v2_data = next((m2[c] for c in m2 if not cas.isdisjoint(c)), None)
                v3_data = next((m3[c] for c in m3 if not cas.isdisjoint(c)), None) if m3 is not None else None

                v1 = v1_data['v'] if v1_data else "누락"
                v2 = v2_data['v'] if v2_data else "누락"
                v3 = (v3_data['v'] if v3_data else "누락") if m3 is not None else None

                name = (v1_data or v2_data or v3_data)['n']

                match = True
                compare_vals = [v for v in [v1, v2, v3] if v is not None]
                
                if "누락" in compare_vals:
                    match = False
                else:
                    it = iter(compare_vals)
                    first = next(it)
                    if not all(abs(first - rest) < 0.0001 for rest in it):
                        match = False

                if not match: mismatch += 1
                
                row_data = {"번호": len(rows)+1, "CAS": ", ".join(list(cas)), "물질명": name, labels[0]: v1, labels[1]: v2}
                if m3 is not None: row_data[labels[2]] = v3
                row_data["상태"] = "✅" if match else "❌"
                rows.append(row_data)

            def get_sum(df_rows, key):
                return sum([r[key] for r in df_rows if isinstance(r[key], (int, float))])
            
            t_a, t_b = get_sum(rows, labels[0]), get_sum(rows, labels[1])
            total_match = abs(t_a - t_b) < 0.0001
            total_row = {"번호": "Total", "CAS": "-", "물질명": "합계", labels[0]: round(t_a, 6), labels[1]: round(t_b, 6)}
            if m3 is not None:
                t_c = get_sum(rows, labels[2])
                total_row[labels[2]] = round(t_c, 6)
                if abs(t_a - t_c) > 0.0001: total_match = False
            total_row["상태"] = "✅" if total_match else "❌"
            rows.append(total_row)

            # 결과 표 출력
            st.expander(f"{'✅' if mismatch == 0 else '❌'} [{idx+1}번] {display_p_name}").dataframe(
                pd.DataFrame(rows).astype(str),
                use_container_width=True, 
                hide_index=True,
                column_config={
                    "CAS": st.column_config.TextColumn("CAS", width="medium", help="마우스를 올리면 전체 CAS 번호가 보입니다.")
                }
            )
            
        except Exception as e:
            st.error(f"{idx+1}번 처리 오류: {e}")
else:
    st.info("검토할 파일들을 모두 업로드해 주세요.")
